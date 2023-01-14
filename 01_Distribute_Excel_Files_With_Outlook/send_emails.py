import os
import openpyxl
import win32com.client as win32


# Get the current working directory
cwd = os.getcwd()

# Load the Excel workbook
workbook = openpyxl.load_workbook(os.path.join(cwd, "Financial_Data.xlsx"))

# Select the sheet
sheet = workbook["Email_List"]

# Get the Outlook application object
outlook = win32.Dispatch('outlook.application')

# Iterate through the rows in the sheet
for i in range(2, sheet.max_row + 1):

    # Get the attachment file name
    attachment = sheet.cell(row=i, column=1).value
    attachment_path = os.path.join(cwd, "Attachments", attachment)
    if not os.path.exists(attachment_path):
        print(f"Attachment {attachment} does not exist")
        continue

    # Get the recipient name
    recipient_name = sheet.cell(row=i, column=2).value

    # Get the recipient email address
    recipient_email = sheet.cell(row=i, column=3).value

    # Get the CC email address
    cc_email = sheet.cell(row=i, column=4).value

    # Create a new email
    mail = outlook.CreateItem(0)

    # Set the recipient and CC email addresses
    mail.To = recipient_email
    mail.CC = cc_email

    # Set the email subject
    mail.Subject = f"Financial Data: {attachment}"

    # Set the email text
    mail.Body = f"Dear {recipient_name},\n\nPlease find the attached financial data for {attachment}.\n\nBest regards,\nYour Name"

    # Add the attachment
    mail.Attachments.Add(attachment_path)

    # Open the email in Outlook
    mail.Display()
    
# close all opened objects
workbook.close()
